"""
Script 5: Run Simulation for All 4 Configs
============================================
Reads:
  db/signals_C1.parquet .. db/signals_C4.parquet   (accumulated per-config signals)
  db/eq_data.parquet                                (OHLCV price history)
  config/params.json                                (configs + simulation params)

Outputs (one per config):
  output/YYYY-MM/C1_Picks_YYYYMMDD_HHMMSS.xlsx
  output/YYYY-MM/C2_Picks_YYYYMMDD_HHMMSS.xlsx
  output/YYYY-MM/C3_Picks_YYYYMMDD_HHMMSS.xlsx
  output/YYYY-MM/C4_Picks_YYYYMMDD_HHMMSS.xlsx

Consolidated (all configs as separate sheets in one file):
  output/YYYY-MM/Consolidated_Picks_YYYYMMDD_HHMMSS.xlsx

Each Excel has sheets:
  Pickse      - per-signal simulation result (same format as QuickRun_Picks)
  MarketData  - latest OHLCV for all EQ symbols
  BuyHistory  - OHLCV history for bought symbols

SIMULATION RULES (same as nse-trading-simulation quick mode):
  - Entry  : D+1 low <= signal_close -> buy at signal_close
  - Additional buys: when buy_count < max_buys AND buy_level >= stop_price
  - Same-day buy and sell NOT allowed
  - Invalid (1): signal date is last available date (no D+1 data)
  - Invalid (2): >10 consecutive calendar days gap
  - Invalid (3): stale data (last date > 10 days ago)
  - Pending: buy not triggered yet, within pending_window_days
  - Expired: buy not triggered, beyond pending_window_days
  - FE-MD  : market_days >= max_duration
  - FE-CD  : calendar_days >= force_exit_calendar_days
"""

import os
import json
import pandas as pd
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG_FILE = "config/params.json"
EQ_FILE     = "db/eq_data.parquet"
OUTPUT_DIR  = "output"
DB_DIR      = "db"

NAVY  = "00203864"
WHITE = "00FFFFFF"
LIGHT = "00F2F2F2"

FMT_DATE  = "DD-MM-YYYY"
FMT_PRICE = "0.00"
FMT_PCT   = "0.00"
FMT_INT   = "General"

OHLCV_COLS = ["SYMBOL", "DATE1", "PREV_CLOSE", "OPEN_PRICE",
              "HIGH_PRICE", "LOW_PRICE", "CLOSE_PRICE"]

IST = timezone(timedelta(hours=5, minutes=30))


# ─── CONFIG ───────────────────────────────────────────────────────────────────

def load_config():
    with open(CONFIG_FILE) as f:
        cfg = json.load(f)
    # Defaults for simulation params (can be overridden in params.json under "simulation")
    sim_cfg = cfg.get("simulation", {})
    defaults = {
        "investment_per_buy":        10000,
        "pending_window_days":       7,
        "force_exit_calendar_days":  90,
    }
    for k, v in defaults.items():
        sim_cfg.setdefault(k, v)
    cfg["simulation"] = sim_cfg
    return cfg


# ─── PRICE DATA ───────────────────────────────────────────────────────────────

def build_price_dict(eq_df):
    print("Building price dictionary...")
    price_dict = {}
    for sym, grp in eq_df.groupby("SYMBOL"):
        grp    = grp.sort_values("DATE1").reset_index(drop=True)
        dates  = grp["DATE1"].values
        closes = grp["CLOSE_PRICE"].values.astype(float)
        highs  = grp["HIGH_PRICE"].values.astype(float)
        lows   = grp["LOW_PRICE"].values.astype(float)
        opens  = grp["OPEN_PRICE"].values.astype(float)
        prevs  = grp["PREV_CLOSE"].values.astype(float)
        day_map = {pd.Timestamp(d): i for i, d in enumerate(dates)}
        price_dict[sym] = {
            "dates":   dates,
            "closes":  closes,
            "highs":   highs,
            "lows":    lows,
            "opens":   opens,
            "prevs":   prevs,
            "day_map": day_map,
        }
    print(f"  Price dict built for {len(price_dict):,} symbols")
    return price_dict


def build_latest_market_data(eq_df):
    print("Extracting latest market data...")
    latest_idx = eq_df.groupby("SYMBOL")["DATE1"].idxmax()
    latest_df  = eq_df.loc[latest_idx].sort_values("SYMBOL").reset_index(drop=True)
    rows = []
    for row in latest_df.itertuples(index=False):
        rows.append([
            str(row.SYMBOL),
            row.DATE1.to_pydatetime().replace(tzinfo=None),
            round(float(row.PREV_CLOSE),  2),
            round(float(row.OPEN_PRICE),  2),
            round(float(row.HIGH_PRICE),  2),
            round(float(row.LOW_PRICE),   2),
            round(float(row.CLOSE_PRICE), 2),
        ])
    print(f"  Market data: {len(rows):,} symbols")
    return rows


def build_buy_history(eq_df, bought_ranges):
    if not bought_ranges:
        return []
    print(f"Extracting buy history for {len(bought_ranges):,} stocks...")
    sym_groups = {sym: grp for sym, grp in eq_df.groupby("SYMBOL")}
    rows = []
    for sym, start_ts, end_ts in bought_ranges:
        if sym not in sym_groups:
            continue
        grp = sym_groups[sym]
        sub = grp[(grp["DATE1"] >= start_ts) & (grp["DATE1"] <= end_ts)]
        if sub.empty:
            continue
        for row in sub.sort_values("DATE1").itertuples(index=False):
            rows.append([
                sym,
                row.DATE1.to_pydatetime().replace(tzinfo=None),
                round(float(row.PREV_CLOSE),  2),
                round(float(row.OPEN_PRICE),  2),
                round(float(row.HIGH_PRICE),  2),
                round(float(row.LOW_PRICE),   2),
                round(float(row.CLOSE_PRICE), 2),
            ])
    print(f"  Buy history rows: {len(rows):,}")
    return rows


# ─── DURATION GROUP ───────────────────────────────────────────────────────────

def duration_group(market_days):
    if market_days is None:
        return None
    if market_days > 50:
        return "50+"
    return ((market_days - 1) // 5 + 1) * 5


# ─── SIMULATION ───────────────────────────────────────────────────────────────

def simulate_trade_detailed(sym, signal_date, signal_close, price_dict,
                             max_buys, buy_drop, target_pct, stoploss_pct,
                             max_duration, investment_per_buy,
                             force_exit_calendar_days, pending_window_days,
                             global_last_date=None):
    invalid = {
        "order": "Invalid", "status": None, "action": "Skip",
        "buy_count": None, "avg_buy_price": None, "total_qty": None,
        "total_investment": None, "target_price": None, "stop_price": None,
        "first_buy_date": None, "exit_found": False,
        "exit_date": None, "exit_price": None, "exit_type": None,
        "profit": None, "gain_pct": None, "market_days": None,
        "result_str": "Invalid: No market data found",
        "duration_group": None, "buy_chance": None,
        "sold_prev_close": None, "sold_open": None,
        "sold_high": None, "sold_low": None, "sold_close": None,
        "buys": [], "had_buy_chance": False,
    }

    if sym not in price_dict:
        return invalid

    pd_data = price_dict[sym]
    dates   = pd_data["dates"]
    closes  = pd_data["closes"]
    highs   = pd_data["highs"]
    lows    = pd_data["lows"]
    opens   = pd_data["opens"]
    prevs   = pd_data["prevs"]
    day_map = pd_data["day_map"]

    sig_ts    = pd.Timestamp(signal_date)
    if sig_ts not in day_map:
        return invalid

    start_idx = day_map[sig_ts]
    last_idx  = len(dates) - 1

    # INVALID 1: signal on last available date — no D+1 data
    if start_idx >= last_idx:
        inv = dict(invalid); inv["result_str"] = "Invalid: No data on signal date"
        return inv

    # INVALID 2: >10 calendar day gap to next data
    next_avail = pd.Timestamp(dates[start_idx + 1])
    if (next_avail - sig_ts).days > 10:
        inv = dict(invalid); inv["result_str"] = "Invalid: Data gap detected"
        return inv

    # INVALID 3: stale data (stock stopped trading)
    last_avail_ts = pd.Timestamp(dates[last_idx])
    today_ts      = pd.Timestamp("today").normalize()
    if (today_ts - last_avail_ts).days > 10:
        inv = dict(invalid); inv["result_str"] = "Invalid: Stale data"
        return inv

    stop_price   = round(signal_close * (1 - stoploss_pct), 2)
    target_price = round(signal_close * (1 + target_pct),  2)

    buy_count        = 0
    avg_buy_price    = 0.0
    total_qty        = 0
    total_investment = 0.0
    first_buy_date   = None
    buy_day_indices  = set()
    market_days      = 0
    exit_found       = False
    exit_type        = None
    exit_price       = 0.0
    exit_date        = None
    exit_idx         = -1
    buys             = []
    had_buy_chance   = False

    prev_date_ts = pd.Timestamp(dates[start_idx])
    for i in range(start_idx + 1, len(dates)):
        curr_ts  = pd.Timestamp(dates[i])

        # Mid-period data gap
        if buy_count > 0 and (curr_ts - prev_date_ts).days > 10:
            return {**invalid, "order": "Invalid", "result_str": "Invalid: Data gap detected"}
        prev_date_ts = curr_ts

        low_px   = float(lows[i])
        high_px  = float(highs[i])
        close_px = float(closes[i])
        open_px  = float(opens[i])
        prev_px  = float(prevs[i])

        if buy_count == 0:
            if low_px <= signal_close:
                buy_count      = 1
                first_buy_date = curr_ts
                buy_day_indices.add(i)
                qty = int(investment_per_buy / signal_close)
                if qty < 1: qty = 1
                total_qty        = qty
                total_investment = signal_close * qty
                avg_buy_price    = round(total_investment / total_qty, 2)
                target_price     = round(avg_buy_price * (1 + target_pct), 2)
                buys.append({
                    "date": curr_ts, "prev_close": prev_px,
                    "open": open_px, "high": high_px,
                    "low": low_px, "close": close_px,
                })
        else:
            is_buy_day = i in buy_day_indices
            if not is_buy_day:
                market_days += 1

            cal_days = (curr_ts - first_buy_date).days if first_buy_date else 0

            if not is_buy_day:
                if low_px <= stop_price:
                    exit_found = True; exit_price = stop_price
                    exit_type  = "Stoploss Triggered"; exit_date = curr_ts; exit_idx = i; break
                if high_px >= target_price:
                    exit_found = True; exit_price = target_price
                    exit_type  = "Target Achieved"; exit_date = curr_ts; exit_idx = i; break
                if market_days >= max_duration:
                    exit_found = True; exit_price = round(close_px, 2)
                    exit_type  = "Force Exit - Market Days"; exit_date = curr_ts; exit_idx = i; break
                if cal_days >= force_exit_calendar_days:
                    exit_found = True; exit_price = round(close_px, 2)
                    exit_type  = "Force Exit - Calendar Days"; exit_date = curr_ts; exit_idx = i; break

            if not exit_found and not is_buy_day and buy_count < max_buys:
                buy_level = round(avg_buy_price * (1 - buy_drop), 2)
                if low_px <= buy_level:
                    had_buy_chance = True
                    if buy_level >= stop_price:
                        buy_count        += 1
                        buy_day_indices.add(i)
                        qty = int(investment_per_buy / buy_level)
                        if qty < 1: qty = 1
                        total_qty        += qty
                        total_investment += buy_level * qty
                        avg_buy_price     = round(total_investment / total_qty, 2)
                        target_price      = round(avg_buy_price * (1 + target_pct), 2)
                        buys.append({
                            "date": curr_ts, "prev_close": prev_px,
                            "open": open_px, "high": high_px,
                            "low": low_px, "close": close_px,
                        })

    # No buy triggered
    if buy_count == 0:
        ref_ts    = pd.Timestamp(global_last_date) if global_last_date else pd.Timestamp(dates[last_idx])
        days_from = (ref_ts - sig_ts).days
        if days_from <= pending_window_days:
            order = "Pending"; action = "Buy"
        else:
            order = "Expired"; action = "Skip"
        return {
            "order": order, "status": "Not Triggered" if order == "Pending" else "Expired",
            "action": action, "buy_count": 0, "avg_buy_price": None, "total_qty": None,
            "total_investment": None, "target_price": target_price, "stop_price": stop_price,
            "first_buy_date": None, "exit_found": False,
            "exit_date": None, "exit_price": None, "exit_type": None,
            "profit": None, "gain_pct": None, "market_days": None,
            "result_str": None, "duration_group": None, "buy_chance": None,
            "sold_prev_close": None, "sold_open": None,
            "sold_high": None, "sold_low": None, "sold_close": None,
            "buys": [], "had_buy_chance": False,
        }

    # Still open
    if not exit_found:
        return {
            "order": "Executed", "status": "Open", "action": "Hold",
            "buy_count": buy_count, "avg_buy_price": avg_buy_price,
            "total_qty": total_qty, "total_investment": round(total_investment, 2),
            "target_price": target_price, "stop_price": stop_price,
            "first_buy_date": first_buy_date, "exit_found": False,
            "exit_date": None, "exit_price": None, "exit_type": None,
            "profit": None, "gain_pct": None, "market_days": market_days,
            "result_str": None, "duration_group": duration_group(market_days),
            "buy_chance": "Buy Chance" if had_buy_chance else None,
            "sold_prev_close": None, "sold_open": None,
            "sold_high": None, "sold_low": None, "sold_close": None,
            "buys": buys, "had_buy_chance": had_buy_chance,
        }

    # Closed
    profit   = round((exit_price - avg_buy_price) * total_qty, 2)
    gain_pct = round(((exit_price - avg_buy_price) / avg_buy_price) * 100, 2) if avg_buy_price > 0 else 0.0

    if   exit_type == "Target Achieved":           result_str = "Profit-TGT"
    elif exit_type == "Stoploss Triggered":         result_str = "Loss-SL"
    elif exit_type == "Force Exit - Market Days":   result_str = ("Profit" if profit >= 0 else "Loss") + "-FE-MD"
    elif exit_type == "Force Exit - Calendar Days": result_str = ("Profit" if profit >= 0 else "Loss") + "-FE-CD"
    else:                                           result_str = "Profit" if profit >= 0 else "Loss"

    sold_prev_close = sold_open = sold_high = sold_low = sold_close = None
    if exit_idx >= 0:
        sold_prev_close = round(float(prevs[exit_idx]),  2)
        sold_open       = round(float(opens[exit_idx]),  2)
        sold_high       = round(float(highs[exit_idx]),  2)
        sold_low        = round(float(lows[exit_idx]),   2)
        sold_close      = round(float(closes[exit_idx]), 2)

    return {
        "order": "Executed", "status": "Closed", "action": "Exit",
        "buy_count": buy_count, "avg_buy_price": avg_buy_price,
        "total_qty": total_qty, "total_investment": round(total_investment, 2),
        "target_price": target_price, "stop_price": stop_price,
        "first_buy_date": first_buy_date, "exit_found": True,
        "exit_date": exit_date, "exit_price": exit_price, "exit_type": exit_type,
        "profit": profit, "gain_pct": gain_pct, "market_days": market_days,
        "result_str": result_str, "duration_group": duration_group(market_days),
        "buy_chance": "Buy Chance" if had_buy_chance else None,
        "sold_prev_close": sold_prev_close, "sold_open": sold_open,
        "sold_high": sold_high, "sold_low": sold_low, "sold_close": sold_close,
        "buys": buys, "had_buy_chance": had_buy_chance,
    }


# ─── PICKS SHEET HELPERS ─────────────────────────────────────────────────────

def get_picks_columns(max_buys):
    cols = [
        "1DChange%", "StockName", "5DLow%", "5DLowPrice", "RecentLTP",
        "BuyDate", "BuyClPrice", "5DLowDate", "TodayDate",
        "BuyCount", "AvgBuyPrice", "TotalQty", "TargetPrice", "StoplossPrice",
        "TotalInvestment", "Order", "Status", "Duration", "DurationGroup",
        "Profits", "GainLoss%", "Result", "ExitType", "Action", "BuyChance",
        "SoldDate", "SoldPrice", "SoldPrevClose", "SoldOpen",
        "SoldHigh", "SoldLow", "SoldClose",
    ]
    for b in range(max_buys):
        cols += [
            f"B{b}_BoughtDate", f"B{b}_PrevClose", f"B{b}_Open",
            f"B{b}_High",       f"B{b}_Low",        f"B{b}_Close",
        ]
    return cols


def _col_fmt(col_name):
    if "Date" in col_name or col_name.endswith("Date"):
        return FMT_DATE
    if col_name in ("1DChange%", "5DLow%", "GainLoss%"):
        return FMT_PCT
    if col_name in ("BuyCount", "TotalQty", "Duration", "DurationGroup", "BuyChance",
                    "Order", "Status", "Result", "ExitType", "Action", "StockName", "WinRate"):
        return FMT_INT
    if col_name in ("5DLowPrice", "RecentLTP", "BuyClPrice", "AvgBuyPrice",
                    "TargetPrice", "StoplossPrice", "TotalInvestment",
                    "Profits", "SoldPrice", "SoldPrevClose", "SoldOpen",
                    "SoldHigh", "SoldLow", "SoldClose"):
        return FMT_PRICE
    if "_PrevClose" in col_name or "_Open" in col_name or "_High" in col_name \
       or "_Low" in col_name or "_Close" in col_name:
        return FMT_PRICE
    return FMT_INT


def _to_dt(ts):
    if ts is None:
        return None
    try:
        t = pd.Timestamp(ts)
        if pd.isna(t):
            return None
        return t.to_pydatetime().replace(tzinfo=None)
    except Exception:
        return None


def build_picks_row(sig, sim, price_dict, max_buys, last_data_date):
    """
    Build one picks-sheet row.
    sig columns: SYMBOL, SIGNAL_DATE, SIGNAL_CLOSE, MIN_5D_LOW, MIN_5D_DATE,
                 PCT_FROM_LOW (%), PCT_FROM_ATH (%), PCT_1D_CHANGE (%)
    """
    sym   = str(sig["SYMBOL"])
    order = sim["order"]

    recent_ltp = None
    if sym in price_dict:
        recent_ltp = round(float(price_dict[sym]["closes"][-1]), 2)
    if recent_ltp is None:
        recent_ltp = round(float(sig["SIGNAL_CLOSE"]), 2)

    buy_date     = _to_dt(sig["SIGNAL_DATE"])
    min5d_date   = _to_dt(sig.get("MIN_5D_DATE"))
    today_date   = last_data_date if order != "Invalid" else None
    sold_date    = _to_dt(sim["exit_date"]) if sim.get("exit_found") else None
    buy_cl_price = round(float(sig["SIGNAL_CLOSE"]), 2)

    buy_chance_val = None
    if order != "Invalid" and recent_ltp is not None and recent_ltp <= buy_cl_price:
        buy_chance_val = "Buy Chance"

    profit   = sim.get("profit")
    gain_pct = sim.get("gain_pct")
    if order == "Executed" and sim.get("status") == "Open":
        avg_buy = sim.get("avg_buy_price")
        qty     = sim.get("total_qty")
        if avg_buy and avg_buy > 0 and recent_ltp and qty:
            profit   = round((recent_ltp - avg_buy) * qty, 2)
            gain_pct = round(((recent_ltp - avg_buy) / avg_buy) * 100, 2)

    # PCT values already stored as percent (e.g. -7.5) — use directly
    row = [
        round(float(sig.get("PCT_1D_CHANGE", 0)), 2),   # 1DChange%
        sym,                                              # StockName
        round(float(sig.get("PCT_FROM_LOW", 0)), 2),    # 5DLow%
        round(float(sig.get("MIN_5D_LOW") or 0), 2),   # 5DLowPrice
        recent_ltp,                                       # RecentLTP
        buy_date,                                         # BuyDate
        buy_cl_price,                                     # BuyClPrice
        min5d_date,                                       # 5DLowDate
        today_date,                                       # TodayDate
        sim.get("buy_count"),                             # BuyCount
        sim.get("avg_buy_price"),                         # AvgBuyPrice
        sim.get("total_qty"),                             # TotalQty
        sim.get("target_price"),                          # TargetPrice
        sim.get("stop_price"),                            # StoplossPrice
        sim.get("total_investment"),                      # TotalInvestment
        sim.get("order"),                                 # Order
        sim.get("status"),                                # Status
        sim.get("market_days"),                           # Duration
        sim.get("duration_group"),                        # DurationGroup
        profit,                                           # Profits
        gain_pct,                                         # GainLoss%
        sim.get("result_str"),                            # Result
        sim.get("exit_type"),                             # ExitType
        sim.get("action"),                                # Action
        buy_chance_val,                                   # BuyChance
        sold_date,                                        # SoldDate
        sim.get("exit_price"),                            # SoldPrice
        sim.get("sold_prev_close"),                       # SoldPrevClose
        sim.get("sold_open"),                             # SoldOpen
        sim.get("sold_high"),                             # SoldHigh
        sim.get("sold_low"),                              # SoldLow
        sim.get("sold_close"),                            # SoldClose
    ]

    buys = sim.get("buys", [])
    for b in range(max_buys):
        if b < len(buys):
            bd = buys[b]
            row += [
                _to_dt(bd["date"]),
                round(float(bd["prev_close"]), 2),
                round(float(bd["open"]),       2),
                round(float(bd["high"]),       2),
                round(float(bd["low"]),        2),
                round(float(bd["close"]),      2),
            ]
        else:
            row += [None, None, None, None, None, None]

    return row


# ─── EXCEL WRITERS ────────────────────────────────────────────────────────────

def _style_ohlcv_header(ws):
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col_name in enumerate(OHLCV_COLS, 1):
        cell           = ws.cell(row=1, column=ci)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
        if col_name == "DATE1":
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=ci).number_format = FMT_DATE
        elif col_name != "SYMBOL":
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=ci).number_format = FMT_PRICE
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 12
    ws.freeze_panes = ws["A2"]


def write_picks_to_sheet(ws, rows, columns):
    """Write picks data to an existing worksheet (used for consolidated Excel)."""
    thin   = Side(style="thin", color="BFBFBF")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols  = len(columns)
    col_fmts = [_col_fmt(c) for c in columns]

    ws.append(columns)
    for col_idx, col_name in enumerate(columns, 1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
    ws.row_dimensions[1].height = 18

    for r_idx, row in enumerate(rows):
        ws.append(row)
        fill_color = LIGHT if (r_idx % 2 == 0) else "00FFFFFF"
        actual_row = r_idx + 2
        for col_idx in range(1, ncols + 1):
            cell               = ws.cell(row=actual_row, column=col_idx)
            cell.fill          = PatternFill("solid", fgColor=fill_color)
            cell.border        = bdr
            cell.alignment     = Alignment(horizontal="center")
            cell.number_format = col_fmts[col_idx - 1]

    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = (ws.cell(row=1, column=1).coordinate + ":" +
                          ws.cell(row=1, column=ncols).coordinate)
    for col_idx, col_name in enumerate(columns, 1):
        w = 13
        if col_name == "StockName": w = 16
        elif "Date" in col_name:    w = 13
        elif col_name in ("Result", "ExitType"): w = 22
        elif col_name in ("Order", "Status", "Action"): w = 14
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def write_picks_excel(rows, columns, out_path, market_data=None, buy_history=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pickse"
    write_picks_to_sheet(ws, rows, columns)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    if market_data:
        ws_md = wb.create_sheet("MarketData")
        ws_md.append(OHLCV_COLS)
        for r in market_data:
            ws_md.append(r)
        _style_ohlcv_header(ws_md)

    if buy_history:
        ws_bh = wb.create_sheet("BuyHistory")
        ws_bh.append(OHLCV_COLS)
        for r in buy_history:
            ws_bh.append(r)
        _style_ohlcv_header(ws_bh)

    wb.save(out_path)
    print(f"  ✅ Saved {out_path} ({len(rows)} rows)")


def write_consolidated_excel(config_data, out_path, market_data=None):
    """
    config_data: list of (config_id, columns, rows)
    Creates one sheet per config in one Excel file.
    """
    wb = Workbook()
    first = True
    for cid, columns, rows in config_data:
        if first:
            ws = wb.active
            ws.title = f"{cid}_Picks"
            first = False
        else:
            ws = wb.create_sheet(f"{cid}_Picks")
        write_picks_to_sheet(ws, rows, columns)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0

    if market_data:
        ws_md = wb.create_sheet("MarketData")
        ws_md.append(OHLCV_COLS)
        for r in market_data:
            ws_md.append(r)
        _style_ohlcv_header(ws_md)

    wb.save(out_path)
    print(f"  ✅ Saved consolidated {out_path}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not os.path.exists(CONFIG_FILE):
        print(f"❌ Missing: {CONFIG_FILE}")
        raise SystemExit(1)
    if not os.path.exists(EQ_FILE):
        print(f"❌ Missing: {EQ_FILE}")
        raise SystemExit(1)

    cfg      = load_config()
    configs  = cfg["configs"]
    sim_cfg  = cfg["simulation"]

    investment_per_buy       = sim_cfg["investment_per_buy"]
    pending_window_days      = sim_cfg["pending_window_days"]
    force_exit_calendar_days = sim_cfg["force_exit_calendar_days"]

    print(f"Simulation params: invest={investment_per_buy}, "
          f"pending_window={pending_window_days}d, fe_cal={force_exit_calendar_days}d")

    # ── Load EQ data ──────────────────────────────────────────────────────────
    print("\nLoading EQ data...")
    eq_df = pd.read_parquet(EQ_FILE, columns=[
        "SYMBOL", "DATE1", "PREV_CLOSE", "OPEN_PRICE",
        "HIGH_PRICE", "LOW_PRICE", "CLOSE_PRICE"
    ])
    eq_df["DATE1"] = pd.to_datetime(eq_df["DATE1"])
    eq_df.sort_values(["SYMBOL", "DATE1"], inplace=True)

    global_last_date = eq_df["DATE1"].max()
    print(f"Global last date: {global_last_date.date()}")

    price_dict   = build_price_dict(eq_df)
    market_data  = build_latest_market_data(eq_df)

    ts_str   = datetime.now(tz=IST).strftime("%Y%m%d_%H%M%S")
    month_dir = os.path.join(OUTPUT_DIR, global_last_date.strftime("%Y-%m"))
    os.makedirs(month_dir, exist_ok=True)

    consolidated_data = []  # list of (cid, columns, rows)

    # ── Run simulation for each config ────────────────────────────────────────
    for c in configs:
        cid          = c["id"]
        max_buys     = c["max_buys"]
        buy_drop     = c["buy_drop"]
        target_pct   = c["target"]
        stoploss_pct = c["stoploss"]
        max_duration = c["max_duration"]

        signals_path = os.path.join(DB_DIR, f"signals_{cid}.parquet")
        if not os.path.exists(signals_path):
            print(f"\n⚠️  {cid}: no signals file found ({signals_path}) — skipping")
            continue

        print(f"\n{'─'*60}")
        print(f"Config {cid}: max_buys={max_buys}, buy_drop={buy_drop}, "
              f"target={target_pct}, sl={stoploss_pct}, max_dur={max_duration}")

        sig_df = pd.read_parquet(signals_path)
        sig_df["SIGNAL_DATE"] = pd.to_datetime(sig_df["SIGNAL_DATE"])
        # Sort by date then symbol
        sig_df = sig_df.sort_values(["SIGNAL_DATE", "SYMBOL"]).reset_index(drop=True)
        print(f"  Signals loaded: {len(sig_df):,} rows ({sig_df['SIGNAL_DATE'].min().date()} → {sig_df['SIGNAL_DATE'].max().date()})")

        columns  = get_picks_columns(max_buys)
        rows     = []
        bought_ranges = []

        for idx, sig in sig_df.iterrows():
            sym          = str(sig["SYMBOL"])
            signal_date  = sig["SIGNAL_DATE"]
            signal_close = float(sig["SIGNAL_CLOSE"])

            sim = simulate_trade_detailed(
                sym, signal_date, signal_close, price_dict,
                max_buys=max_buys, buy_drop=buy_drop,
                target_pct=target_pct, stoploss_pct=stoploss_pct,
                max_duration=max_duration,
                investment_per_buy=investment_per_buy,
                force_exit_calendar_days=force_exit_calendar_days,
                pending_window_days=pending_window_days,
                global_last_date=global_last_date,
            )

            row = build_picks_row(sig, sim, price_dict, max_buys,
                                  global_last_date.to_pydatetime().replace(tzinfo=None))
            rows.append(row)

            # Track bought ranges for BuyHistory sheet
            if sim.get("first_buy_date") is not None:
                start_ts = pd.Timestamp(signal_date) - pd.Timedelta(days=7)
                end_ts   = pd.Timestamp(sim["exit_date"]) if sim.get("exit_date") else global_last_date
                bought_ranges.append((sym, start_ts, end_ts))

        buy_history = build_buy_history(eq_df, bought_ranges)

        # Summary
        orders  = [r[15] for r in rows]  # col index 15 = Order
        statuses = [r[16] for r in rows]  # col index 16 = Status
        executed = sum(1 for o in orders if o == "Executed")
        open_c   = sum(1 for s in statuses if s == "Open")
        closed_c = sum(1 for s in statuses if s == "Closed")
        pending  = sum(1 for o in orders if o == "Pending")
        expired  = sum(1 for o in orders if o == "Expired")
        invalid  = sum(1 for o in orders if o == "Invalid")
        print(f"  Results: Executed={executed} (Open={open_c}, Closed={closed_c}) "
              f"Pending={pending} Expired={expired} Invalid={invalid}")

        # Per-config Excel
        out_path = os.path.join(month_dir, f"{cid}_Picks_{ts_str}.xlsx")
        write_picks_excel(rows, columns, out_path,
                          market_data=market_data, buy_history=buy_history)

        consolidated_data.append((cid, columns, rows))

    # ── Consolidated Excel ────────────────────────────────────────────────────
    if consolidated_data:
        cons_path = os.path.join(month_dir, f"Consolidated_Picks_{ts_str}.xlsx")
        write_consolidated_excel(consolidated_data, cons_path, market_data=market_data)
    else:
        print("\n⚠️  No config data to consolidate — no signal parquets found yet")
        print("   Run the Bootstrap workflow first, then Daily Signals.")

    print(f"\n{'='*60}")
    print(f"Simulation complete — {datetime.now(tz=IST).strftime('%d-%b-%Y %H:%M IST')}")
    print(f"Output folder: {month_dir}")


if __name__ == "__main__":
    main()
